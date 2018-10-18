# -*- coding:utf8 -*-
import sys

reload(sys)
sys.setdefaultencoding('utf-8')

import xlrd
from openpyxl import Workbook
from RowItem import RowItem

def find_key_index(key, data):
    keyIndex = data.find(key + ':')
    if keyIndex == -1:
        keyIndex = data.find(key + '.')
        if keyIndex == -1:
            keyIndex = data.find(key + '、')
            if keyIndex == -1:
                keyIndex = data.find(key + '-')

    return keyIndex

# 解析文件题干信息
def parse_data(info):
    # print("line info:" + info)

    key_set = info[0:len(info)]

    key_a_index = find_key_index('A', key_set)
    key_b_index = find_key_index('B', key_set)
    key_c_index = find_key_index('C', key_set)
    key_d_index = find_key_index('D', key_set)
    key_e_index = find_key_index('E', key_set)
    key_f_index = find_key_index('F', key_set)

    # 第一个换行默认为题干结束
    question = info[0:key_a_index - 1]
    item = RowItem(question)

    key_a_question = key_set[key_a_index + 2:key_b_index]
    item.keyA = key_a_question

    key_b_question = key_set[key_b_index + 2:key_c_index]
    item.keyB = key_b_question

    key_c_question = key_set[key_c_index + 2:key_d_index]
    item.keyC = key_c_question

    if key_e_index != -1 and key_d_index < key_e_index:
        key_d_question = key_set[key_d_index + 2:key_e_index]
        item.keyD = key_d_question
    elif key_e_index == -1 and key_d_index != -1:
        key_d_question = key_set[key_d_index + 2:]
        item.keyD = key_d_question

    return item


# 准备写文件

outbook = Workbook()

formate_sheet = outbook.create_sheet('format')


# 准备写标题
formate_sheet.cell(1, 1, '题干')
formate_sheet.cell(1, 2, '选项A')
formate_sheet.cell(1, 3, '选项B')
formate_sheet.cell(1, 4, '选项C')
formate_sheet.cell(1, 5, '选项D')
formate_sheet.cell(1, 6, '选项E')
formate_sheet.cell(1, 7, '选项F')
formate_sheet.cell(1, 8, '标准答案')


# read excel

# 1、读取文件
workbook = xlrd.open_workbook(r'librarys.xlsx')

# 2、读取LTE单选sheet数据内容
single_select_sheet = workbook.sheet_by_name('LTE单选')

print single_select_sheet.name, single_select_sheet.ncols, single_select_sheet.nrows
rows_size = single_select_sheet.nrows

# 3、处理表格数据
# 注意:此处下标从2开始，第一行要显示标题
for i in range(2, rows_size, 1):
    print "===========================================i:", i

    # 读取每行题干信息(第二列数据)
    info = single_select_sheet.cell(i, 2).value

    # 对题干数据进行解析处理，生成一个Bean对象
    item = parse_data(info)

    # 打印信息
    print('question:' + item.question)
    print('itemA:' + item.keyA)
    print('itemB:' + item.keyB)
    print('itemC:' + item.keyC)
    print('itemD:' + item.keyD)
    print('itemE:' + item.keyE)
    print('itemF:' + item.keyF)

    # 将题干，A B C D E F 分别写到对应行对对应列(注意:cell方法下标行，列都是从1开始的)
    formate_sheet.cell(i, 1, item.question.encode('utf-8'))
    formate_sheet.cell(i, 2, item.keyA.encode('utf-8'))
    formate_sheet.cell(i, 3, item.keyB.encode('utf-8'))
    formate_sheet.cell(i, 4, item.keyC.encode('utf-8'))
    formate_sheet.cell(i, 5, item.keyD.encode('utf-8'))
    formate_sheet.cell(i, 6, item.keyE.encode('utf-8'))
    formate_sheet.cell(i, 7, item.keyF.encode('utf-8'))

    # 读取参考答案(第三列数据)
    key = single_select_sheet.cell(i, 3).value.encode('utf-8')
    formate_sheet.cell(i, 8, key)

# 将整理的数据写入到新excel表格中
outbook.save('single_select_result.xlsx')









